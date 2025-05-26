import asyncio
import json
import logging
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass, field
from enum import Enum
from datetime import datetime
import numpy as np
import sqlite3
from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import base64

logger = logging.getLogger(__name__)

class DocumentationType(Enum):
    """Control system documentation types"""
    IO_LIST = "io_list"
    INSTRUMENT_INDEX = "instrument_index"
    CONTROL_NARRATIVE = "control_narrative"
    LOOP_DIAGRAM = "loop_diagram"
    CAUSE_EFFECT_MATRIX = "cause_effect_matrix"
    ALARM_SCHEDULE = "alarm_schedule"
    INTERLOCK_MATRIX = "interlock_matrix"
    CABLE_SCHEDULE = "cable_schedule"
    JUNCTION_BOX_SCHEDULE = "junction_box_schedule"
    POWER_CONSUMPTION = "power_consumption"
    NETWORK_ARCHITECTURE = "network_architecture"
    CYBERSECURITY_MATRIX = "cybersecurity_matrix"

class TagNumberingStandard(Enum):
    """Tag numbering standards"""
    ISA_S5_1 = "isa_s5_1"  # ISA-5.1 Instrumentation Symbols
    ISA_S5_3 = "isa_s5_3"  # ISA-5.3 Computer/Human Interface
    ANSI_ISA_S18_1 = "ansi_isa_s18_1"  # Alarm Management
    IEC_62424 = "iec_62424"  # Representation of process control engineering
    CUSTOM = "custom"

class ExportFormat(Enum):
    """Documentation export formats"""
    EXCEL = "excel"
    PDF = "pdf"
    CSV = "csv"
    XML = "xml"
    JSON = "json"
    AUTOCAD = "autocad"
    MICROSTATION = "microstation"

@dataclass
class InstrumentTag:
    """Instrument tag specification"""
    tag_number: str
    description: str
    service: str
    instrument_type: str
    location: str
    manufacturer: str = ""
    model_number: str = ""
    measurement_range: str = ""
    units: str = ""
    accuracy: str = ""
    signal_type: str = ""
    power_supply: str = ""
    environmental_conditions: str = ""
    calibration_frequency: str = ""
    safety_classification: str = ""
    sil_rating: str = ""

@dataclass
class IOPoint:
    """I/O point specification"""
    tag_number: str
    description: str
    io_type: str  # AI, AO, DI, DO
    signal_range: str
    engineering_units: str
    controller_location: str
    terminal_block: str
    wire_number: str
    cable_type: str
    signal_conditioning: str = ""
    alarm_limits: Dict[str, float] = field(default_factory=dict)
    control_function: str = ""

@dataclass
class ControlLoop:
    """Control loop specification"""
    loop_tag: str
    description: str
    controlled_variable: str
    process_variable_tag: str
    controller_tag: str
    final_element_tag: str
    setpoint_range: str
    control_mode: str  # Manual, Auto, Cascade, etc.
    tuning_parameters: Dict[str, float] = field(default_factory=dict)
    interlock_conditions: List[str] = field(default_factory=list)
    alarm_conditions: List[str] = field(default_factory=list)

@dataclass
class CableSpecification:
    """Cable specification"""
    cable_number: str
    cable_type: str
    from_location: str
    to_location: str
    core_count: int
    conductor_size: str
    length: float
    routing: str
    installation_method: str
    environmental_rating: str = ""
    fire_rating: str = ""
    armoring: str = ""

@dataclass
class ControlSystemDocumentation:
    """Complete control system documentation package"""
    system_id: str
    project_name: str
    project_number: str
    revision: str
    date_created: datetime
    engineer_name: str
    instrument_tags: List[InstrumentTag] = field(default_factory=list)
    io_points: List[IOPoint] = field(default_factory=list)
    control_loops: List[ControlLoop] = field(default_factory=list)
    cable_schedule: List[CableSpecification] = field(default_factory=list)
    network_architecture: Dict[str, Any] = field(default_factory=dict)
    power_requirements: Dict[str, Any] = field(default_factory=dict)
    safety_systems: Dict[str, Any] = field(default_factory=dict)

class ControlSystemDocumentationGenerator:
    """Generate comprehensive control system documentation"""
    
    def __init__(self, tag_standard: TagNumberingStandard = TagNumberingStandard.ISA_S5_1):
        self.tag_standard = tag_standard
        self.tag_generator = TagNumberGenerator(tag_standard)
        self.io_analyzer = IOSystemAnalyzer()
        self.document_formatter = DocumentFormatter()
        
    async def generate_complete_documentation(self, 
                                            system_components: List[Dict[str, Any]],
                                            project_info: Dict[str, str]) -> ControlSystemDocumentation:
        """Generate complete control system documentation"""
        
        documentation = ControlSystemDocumentation(
            system_id=project_info.get("system_id", "SYS001"),
            project_name=project_info.get("project_name", "Control System"),
            project_number=project_info.get("project_number", "P2024-001"),
            revision=project_info.get("revision", "A"),
            date_created=datetime.now(),
            engineer_name=project_info.get("engineer", "System Engineer")
        )
        
        # Generate instrument tags
        documentation.instrument_tags = await self._generate_instrument_tags(system_components)
        
        # Generate I/O list
        documentation.io_points = await self._generate_io_list(system_components)
        
        # Generate control loops
        documentation.control_loops = await self._generate_control_loops(system_components)
        
        # Generate cable schedule
        documentation.cable_schedule = await self._generate_cable_schedule(system_components)
        
        # Analyze network architecture
        documentation.network_architecture = await self._analyze_network_architecture(system_components)
        
        # Calculate power requirements
        documentation.power_requirements = await self._calculate_power_requirements(system_components)
        
        # Analyze safety systems
        documentation.safety_systems = await self._analyze_safety_systems(system_components)
        
        return documentation
    
    async def _generate_instrument_tags(self, components: List[Dict[str, Any]]) -> List[InstrumentTag]:
        """Generate instrument tag index"""
        instrument_tags = []
        
        for component in components:
            if self._is_instrument(component):
                tag = self.tag_generator.generate_instrument_tag(component)
                
                instrument_tag = InstrumentTag(
                    tag_number=tag,
                    description=component.get("description", "Instrument"),
                    service=component.get("service", "Process Control"),
                    instrument_type=component.get("category", "sensor"),
                    location=component.get("location", "Field"),
                    manufacturer=component.get("manufacturer", "TBD"),
                    model_number=component.get("model_number", "TBD"),
                    measurement_range=component.get("measurement_range", "TBD"),
                    units=component.get("units", ""),
                    accuracy=component.get("accuracy", "TBD"),
                    signal_type=component.get("signal_type", "4-20mA"),
                    power_supply=component.get("power_supply", "24VDC"),
                    environmental_conditions=component.get("environmental", "Indoor"),
                    calibration_frequency=component.get("calibration", "Annually"),
                    safety_classification=component.get("safety_class", "Normal"),
                    sil_rating=component.get("sil_rating", "")
                )
                
                instrument_tags.append(instrument_tag)
        
        return sorted(instrument_tags, key=lambda x: x.tag_number)
    
    async def _generate_io_list(self, components: List[Dict[str, Any]]) -> List[IOPoint]:
        """Generate comprehensive I/O list"""
        io_points = []
        
        for component in components:
            if self._has_io_points(component):
                points = self.io_analyzer.analyze_component_io(component)
                
                for point in points:
                    io_point = IOPoint(
                        tag_number=point["tag"],
                        description=point["description"],
                        io_type=point["type"],
                        signal_range=point["range"],
                        engineering_units=point["units"],
                        controller_location=point["controller"],
                        terminal_block=point["terminal_block"],
                        wire_number=point["wire_number"],
                        cable_type=point["cable_type"],
                        signal_conditioning=point.get("conditioning", ""),
                        alarm_limits=point.get("alarms", {}),
                        control_function=point.get("function", "")
                    )
                    
                    io_points.append(io_point)
        
        return sorted(io_points, key=lambda x: x.tag_number)
    
    async def _generate_control_loops(self, components: List[Dict[str, Any]]) -> List[ControlLoop]:
        """Generate control loop documentation"""
        control_loops = []
        
        # Identify control loops from components
        loop_components = self._identify_control_loops(components)
        
        for loop_id, loop_data in loop_components.items():
            control_loop = ControlLoop(
                loop_tag=loop_id,
                description=loop_data["description"],
                controlled_variable=loop_data["controlled_variable"],
                process_variable_tag=loop_data["pv_tag"],
                controller_tag=loop_data["controller_tag"],
                final_element_tag=loop_data["final_element_tag"],
                setpoint_range=loop_data["setpoint_range"],
                control_mode=loop_data["control_mode"],
                tuning_parameters=loop_data.get("tuning", {}),
                interlock_conditions=loop_data.get("interlocks", []),
                alarm_conditions=loop_data.get("alarms", [])
            )
            
            control_loops.append(control_loop)
        
        return sorted(control_loops, key=lambda x: x.loop_tag)
    
    async def _generate_cable_schedule(self, components: List[Dict[str, Any]]) -> List[CableSpecification]:
        """Generate cable schedule"""
        cables = []
        cable_counter = 1
        
        for component in components:
            if self._requires_cabling(component):
                cable_specs = self._generate_component_cables(component, cable_counter)
                cables.extend(cable_specs)
                cable_counter += len(cable_specs)
        
        return cables
    
    async def _analyze_network_architecture(self, components: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze control system network architecture"""
        network_devices = [c for c in components if self._is_network_device(c)]
        
        return {
            "network_topology": self._determine_topology(network_devices),
            "communication_protocols": self._analyze_protocols(components),
            "network_segments": self._identify_network_segments(network_devices),
            "bandwidth_requirements": self._calculate_bandwidth(components),
            "redundancy_analysis": self._analyze_redundancy(network_devices),
            "cybersecurity_zones": self._identify_security_zones(network_devices)
        }
    
    async def _calculate_power_requirements(self, components: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate system power requirements"""
        power_analysis = {
            "total_power_consumption": 0,
            "power_by_voltage": {},
            "power_by_area": {},
            "ups_requirements": {},
            "power_distribution": []
        }
        
        for component in components:
            power_data = self._extract_power_data(component)
            if power_data:
                power_analysis["total_power_consumption"] += power_data["consumption"]
                
                voltage = power_data["voltage"]
                if voltage not in power_analysis["power_by_voltage"]:
                    power_analysis["power_by_voltage"][voltage] = 0
                power_analysis["power_by_voltage"][voltage] += power_data["consumption"]
                
                area = component.get("area", "General")
                if area not in power_analysis["power_by_area"]:
                    power_analysis["power_by_area"][area] = 0
                power_analysis["power_by_area"][area] += power_data["consumption"]
        
        # Calculate UPS requirements
        critical_power = sum(
            self._extract_power_data(c)["consumption"] 
            for c in components 
            if c.get("critical", False) and self._extract_power_data(c)
        )
        
        power_analysis["ups_requirements"] = {
            "critical_load": critical_power,
            "recommended_capacity": critical_power * 1.25,  # 25% safety factor
            "backup_time": "30 minutes minimum"
        }
        
        return power_analysis
    
    async def _analyze_safety_systems(self, components: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze safety systems"""
        safety_components = [c for c in components if c.get("safety_related", False)]
        
        return {
            "safety_functions": self._identify_safety_functions(safety_components),
            "sil_analysis": self._analyze_sil_requirements(safety_components),
            "safety_loops": self._identify_safety_loops(safety_components),
            "proof_test_schedule": self._generate_proof_test_schedule(safety_components),
            "safety_lifecycle": self._generate_safety_lifecycle_plan(safety_components)
        }
    
    def _is_instrument(self, component: Dict[str, Any]) -> bool:
        """Check if component is an instrument"""
        instrument_categories = [
            "pressure_transmitter", "temperature_transmitter", "flow_transmitter",
            "level_transmitter", "ph_transmitter", "conductivity_transmitter",
            "gas_detector", "flame_detector", "smoke_detector"
        ]
        return component.get("category") in instrument_categories
    
    def _has_io_points(self, component: Dict[str, Any]) -> bool:
        """Check if component has I/O points"""
        io_categories = [
            "plc", "bas_controller", "vav_controller", "ahu_controller",
            "pressure_transmitter", "temperature_transmitter", "control_valve"
        ]
        return component.get("category") in io_categories
    
    def _requires_cabling(self, component: Dict[str, Any]) -> bool:
        """Check if component requires cabling"""
        return component.get("location") != "wireless" and "transmitter" in component.get("category", "")
    
    def _is_network_device(self, component: Dict[str, Any]) -> bool:
        """Check if component is a network device"""
        network_categories = ["plc", "hmi", "gateway", "switch", "router"]
        return component.get("category") in network_categories
    
    def _identify_control_loops(self, components: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """Identify control loops from components"""
        loops = {}
        
        # Simple loop identification logic
        transmitters = [c for c in components if "transmitter" in c.get("category", "")]
        controllers = [c for c in components if c.get("category") == "plc"]
        valves = [c for c in components if "valve" in c.get("category", "")]
        
        for i, transmitter in enumerate(transmitters):
            loop_id = f"LC-{i+1:03d}"
            loops[loop_id] = {
                "description": f"Control Loop {i+1}",
                "controlled_variable": transmitter.get("measurement_type", "Process Variable"),
                "pv_tag": transmitter.get("tag", f"PT-{i+1:03d}"),
                "controller_tag": controllers[0].get("tag", "PLC-001") if controllers else "TBD",
                "final_element_tag": valves[i].get("tag", f"PV-{i+1:03d}") if i < len(valves) else "TBD",
                "setpoint_range": transmitter.get("measurement_range", "0-100%"),
                "control_mode": "Auto"
            }
        
        return loops
    
    def _generate_component_cables(self, component: Dict[str, Any], start_number: int) -> List[CableSpecification]:
        """Generate cable specifications for a component"""
        cables = []
        
        cable = CableSpecification(
            cable_number=f"C{start_number:04d}",
            cable_type="Instrumentation Cable",
            from_location=component.get("location", "Field"),
            to_location="Control Room",
            core_count=2,
            conductor_size="18 AWG",
            length=100.0,  # Default length
            routing="Cable Tray",
            installation_method="Tray Installation",
            environmental_rating="Indoor",
            fire_rating="LSZH",
            armoring="SWA"
        )
        
        cables.append(cable)
        return cables
    
    def _extract_power_data(self, component: Dict[str, Any]) -> Optional[Dict[str, float]]:
        """Extract power consumption data from component"""
        if "power_consumption" in component:
            return {
                "consumption": component["power_consumption"],
                "voltage": component.get("power_voltage", 24)
            }
        return None

class TagNumberGenerator:
    """Generate instrument tag numbers according to standards"""
    
    def __init__(self, standard: TagNumberingStandard):
        self.standard = standard
        self.tag_counters = {}
    
    def generate_instrument_tag(self, component: Dict[str, Any]) -> str:
        """Generate instrument tag number"""
        if self.standard == TagNumberingStandard.ISA_S5_1:
            return self._generate_isa_s5_1_tag(component)
        else:
            return self._generate_custom_tag(component)
    
    def _generate_isa_s5_1_tag(self, component: Dict[str, Any]) -> str:
        """Generate ISA-5.1 compliant tag number"""
        # ISA-5.1 tag format: [Process Area][First Letter][Subsequent Letters][Loop Number]
        
        category = component.get("category", "sensor")
        process_area = component.get("area", "")
        
        # Determine first letter (measured variable)
        first_letter_map = {
            "pressure_transmitter": "P",
            "temperature_transmitter": "T",
            "flow_transmitter": "F",
            "level_transmitter": "L",
            "conductivity_transmitter": "C",
            "ph_transmitter": "P",
            "gas_detector": "G",
            "flame_detector": "F",
            "control_valve": "F"  # Flow control
        }
        
        first_letter = first_letter_map.get(category, "X")
        
        # Determine subsequent letters (function)
        if "transmitter" in category:
            subsequent_letters = "T"
        elif "valve" in category:
            subsequent_letters = "V"
        elif "indicator" in category:
            subsequent_letters = "I"
        elif "controller" in category:
            subsequent_letters = "C"
        else:
            subsequent_letters = "E"  # Element
        
        # Get loop number
        tag_key = f"{first_letter}{subsequent_letters}"
        if tag_key not in self.tag_counters:
            self.tag_counters[tag_key] = 0
        self.tag_counters[tag_key] += 1
        
        loop_number = f"{self.tag_counters[tag_key]:03d}"
        
        return f"{process_area}{first_letter}{subsequent_letters}-{loop_number}"
    
    def _generate_custom_tag(self, component: Dict[str, Any]) -> str:
        """Generate custom tag number"""
        category = component.get("category", "DEV")
        area = component.get("area", "")
        
        tag_key = category.upper()
        if tag_key not in self.tag_counters:
            self.tag_counters[tag_key] = 0
        self.tag_counters[tag_key] += 1
        
        return f"{area}{tag_key}-{self.tag_counters[tag_key]:03d}"

class IOSystemAnalyzer:
    """Analyze I/O system requirements"""
    
    def analyze_component_io(self, component: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze I/O points for a component"""
        io_points = []
        
        category = component.get("category", "")
        
        if "transmitter" in category:
            # Analog input for process variable
            io_points.append({
                "tag": component.get("tag", "AI-001"),
                "description": f"{component.get('description', 'Process Variable')} - PV",
                "type": "AI",
                "range": "4-20 mA",
                "units": component.get("units", ""),
                "controller": "PLC-001",
                "terminal_block": "TB1",
                "wire_number": "W001",
                "cable_type": "Instrumentation"
            })
        
        elif category == "control_valve":
            # Analog output for valve position
            io_points.append({
                "tag": component.get("tag", "AO-001"),
                "description": f"{component.get('description', 'Control Valve')} - Output",
                "type": "AO",
                "range": "4-20 mA",
                "units": "%",
                "controller": "PLC-001",
                "terminal_block": "TB2",
                "wire_number": "W002",
                "cable_type": "Instrumentation"
            })
            
            # Digital inputs for limit switches
            io_points.extend([
                {
                    "tag": f"{component.get('tag', 'DI')}-FS",
                    "description": f"{component.get('description', 'Valve')} - Fully Open",
                    "type": "DI",
                    "range": "24 VDC",
                    "units": "",
                    "controller": "PLC-001",
                    "terminal_block": "TB3",
                    "wire_number": "W003",
                    "cable_type": "Control"
                },
                {
                    "tag": f"{component.get('tag', 'DI')}-FC",
                    "description": f"{component.get('description', 'Valve')} - Fully Closed",
                    "type": "DI",
                    "range": "24 VDC",
                    "units": "",
                    "controller": "PLC-001",
                    "terminal_block": "TB3",
                    "wire_number": "W004",
                    "cable_type": "Control"
                }
            ])
        
        elif category == "motor":
            # Digital outputs for motor control
            io_points.extend([
                {
                    "tag": f"{component.get('tag', 'DO')}-START",
                    "description": f"{component.get('description', 'Motor')} - Start",
                    "type": "DO",
                    "range": "24 VDC",
                    "units": "",
                    "controller": "PLC-001",
                    "terminal_block": "TB4",
                    "wire_number": "W005",
                    "cable_type": "Control"
                },
                {
                    "tag": f"{component.get('tag', 'DI')}-RUN",
                    "description": f"{component.get('description', 'Motor')} - Running",
                    "type": "DI",
                    "range": "24 VDC",
                    "units": "",
                    "controller": "PLC-001",
                    "terminal_block": "TB4",
                    "wire_number": "W006",
                    "cable_type": "Control"
                }
            ])
        
        return io_points

class DocumentFormatter:
    """Format documentation for export"""
    
    def export_to_excel(self, documentation: ControlSystemDocumentation, file_path: str):
        """Export documentation to Excel format"""
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        self._create_instrument_index_sheet(workbook, documentation)
        self._create_io_list_sheet(workbook, documentation)
        self._create_control_loops_sheet(workbook, documentation)
        self._create_cable_schedule_sheet(workbook, documentation)
        self._create_power_summary_sheet(workbook, documentation)
        
        workbook.save(file_path)
    
    def _create_instrument_index_sheet(self, workbook: Workbook, doc: ControlSystemDocumentation):
        """Create instrument index sheet"""
        sheet = workbook.create_sheet("Instrument Index")
        
        # Headers
        headers = [
            "Tag Number", "Description", "Service", "Type", "Location",
            "Manufacturer", "Model", "Range", "Units", "Accuracy",
            "Signal Type", "Power Supply", "Environment", "Calibration",
            "Safety Class", "SIL Rating"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, tag in enumerate(doc.instrument_tags, 2):
            sheet.cell(row=row, column=1, value=tag.tag_number)
            sheet.cell(row=row, column=2, value=tag.description)
            sheet.cell(row=row, column=3, value=tag.service)
            sheet.cell(row=row, column=4, value=tag.instrument_type)
            sheet.cell(row=row, column=5, value=tag.location)
            sheet.cell(row=row, column=6, value=tag.manufacturer)
            sheet.cell(row=row, column=7, value=tag.model_number)
            sheet.cell(row=row, column=8, value=tag.measurement_range)
            sheet.cell(row=row, column=9, value=tag.units)
            sheet.cell(row=row, column=10, value=tag.accuracy)
            sheet.cell(row=row, column=11, value=tag.signal_type)
            sheet.cell(row=row, column=12, value=tag.power_supply)
            sheet.cell(row=row, column=13, value=tag.environmental_conditions)
            sheet.cell(row=row, column=14, value=tag.calibration_frequency)
            sheet.cell(row=row, column=15, value=tag.safety_classification)
            sheet.cell(row=row, column=16, value=tag.sil_rating)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_io_list_sheet(self, workbook: Workbook, doc: ControlSystemDocumentation):
        """Create I/O list sheet"""
        sheet = workbook.create_sheet("I/O List")
        
        headers = [
            "Tag Number", "Description", "I/O Type", "Signal Range", "Units",
            "Controller", "Terminal Block", "Wire Number", "Cable Type",
            "Conditioning", "Alarm Limits", "Function"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, io_point in enumerate(doc.io_points, 2):
            sheet.cell(row=row, column=1, value=io_point.tag_number)
            sheet.cell(row=row, column=2, value=io_point.description)
            sheet.cell(row=row, column=3, value=io_point.io_type)
            sheet.cell(row=row, column=4, value=io_point.signal_range)
            sheet.cell(row=row, column=5, value=io_point.engineering_units)
            sheet.cell(row=row, column=6, value=io_point.controller_location)
            sheet.cell(row=row, column=7, value=io_point.terminal_block)
            sheet.cell(row=row, column=8, value=io_point.wire_number)
            sheet.cell(row=row, column=9, value=io_point.cable_type)
            sheet.cell(row=row, column=10, value=io_point.signal_conditioning)
            sheet.cell(row=row, column=11, value=str(io_point.alarm_limits))
            sheet.cell(row=row, column=12, value=io_point.control_function)
    
    def _create_control_loops_sheet(self, workbook: Workbook, doc: ControlSystemDocumentation):
        """Create control loops sheet"""
        sheet = workbook.create_sheet("Control Loops")
        
        headers = [
            "Loop Tag", "Description", "Controlled Variable", "PV Tag",
            "Controller Tag", "Final Element", "Setpoint Range", "Control Mode",
            "Tuning Parameters", "Interlocks", "Alarms"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, loop in enumerate(doc.control_loops, 2):
            sheet.cell(row=row, column=1, value=loop.loop_tag)
            sheet.cell(row=row, column=2, value=loop.description)
            sheet.cell(row=row, column=3, value=loop.controlled_variable)
            sheet.cell(row=row, column=4, value=loop.process_variable_tag)
            sheet.cell(row=row, column=5, value=loop.controller_tag)
            sheet.cell(row=row, column=6, value=loop.final_element_tag)
            sheet.cell(row=row, column=7, value=loop.setpoint_range)
            sheet.cell(row=row, column=8, value=loop.control_mode)
            sheet.cell(row=row, column=9, value=str(loop.tuning_parameters))
            sheet.cell(row=row, column=10, value=", ".join(loop.interlock_conditions))
            sheet.cell(row=row, column=11, value=", ".join(loop.alarm_conditions))
    
    def _create_cable_schedule_sheet(self, workbook: Workbook, doc: ControlSystemDocumentation):
        """Create cable schedule sheet"""
        sheet = workbook.create_sheet("Cable Schedule")
        
        headers = [
            "Cable Number", "Cable Type", "From", "To", "Cores",
            "Conductor Size", "Length (m)", "Routing", "Installation",
            "Environmental", "Fire Rating", "Armoring"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, cable in enumerate(doc.cable_schedule, 2):
            sheet.cell(row=row, column=1, value=cable.cable_number)
            sheet.cell(row=row, column=2, value=cable.cable_type)
            sheet.cell(row=row, column=3, value=cable.from_location)
            sheet.cell(row=row, column=4, value=cable.to_location)
            sheet.cell(row=row, column=5, value=cable.core_count)
            sheet.cell(row=row, column=6, value=cable.conductor_size)
            sheet.cell(row=row, column=7, value=cable.length)
            sheet.cell(row=row, column=8, value=cable.routing)
            sheet.cell(row=row, column=9, value=cable.installation_method)
            sheet.cell(row=row, column=10, value=cable.environmental_rating)
            sheet.cell(row=row, column=11, value=cable.fire_rating)
            sheet.cell(row=row, column=12, value=cable.armoring)
    
    def _create_power_summary_sheet(self, workbook: Workbook, doc: ControlSystemDocumentation):
        """Create power requirements summary sheet"""
        sheet = workbook.create_sheet("Power Summary")
        
        # Title
        sheet.cell(row=1, column=1, value="Power Requirements Summary").font = Font(bold=True, size=14)
        
        # Total power
        sheet.cell(row=3, column=1, value="Total Power Consumption:")
        sheet.cell(row=3, column=2, value=f"{doc.power_requirements.get('total_power_consumption', 0)} W")
        
        # Power by voltage
        row = 5
        sheet.cell(row=row, column=1, value="Power by Voltage:").font = Font(bold=True)
        for voltage, power in doc.power_requirements.get("power_by_voltage", {}).items():
            row += 1
            sheet.cell(row=row, column=1, value=f"{voltage}V:")
            sheet.cell(row=row, column=2, value=f"{power} W")
        
        # UPS requirements
        row += 2
        sheet.cell(row=row, column=1, value="UPS Requirements:").font = Font(bold=True)
        ups_req = doc.power_requirements.get("ups_requirements", {})
        for key, value in ups_req.items():
            row += 1
            sheet.cell(row=row, column=1, value=f"{key.replace('_', ' ').title()}:")
            sheet.cell(row=row, column=2, value=str(value))

# Testing functions
async def test_documentation_generation():
    """Test control system documentation generation"""
    generator = ControlSystemDocumentationGenerator()
    
    # Sample system components
    sample_components = [
        {
            "category": "pressure_transmitter",
            "description": "Inlet Pressure",
            "location": "Field",
            "manufacturer": "Rosemount",
            "model_number": "3051S",
            "measurement_range": "0-100 psi",
            "units": "psi",
            "power_consumption": 5,
            "power_voltage": 24,
            "area": "UNIT1"
        },
        {
            "category": "control_valve",
            "description": "Flow Control Valve",
            "location": "Field",
            "manufacturer": "Fisher",
            "model_number": "ED Series",
            "power_consumption": 8,
            "power_voltage": 24,
            "area": "UNIT1"
        },
        {
            "category": "plc",
            "description": "Main Process Controller",
            "location": "Control Room",
            "manufacturer": "Allen-Bradley",
            "model_number": "CompactLogix",
            "power_consumption": 50,
            "power_voltage": 120,
            "area": "CR"
        }
    ]
    
    project_info = {
        "system_id": "SYS001",
        "project_name": "Test Control System",
        "project_number": "P2024-001",
        "revision": "A",
        "engineer": "Test Engineer"
    }
    
    # Generate documentation
    documentation = await generator.generate_complete_documentation(sample_components, project_info)
    
    print(f"Documentation Generated:")
    print(f"Project: {documentation.project_name}")
    print(f"Instrument Tags: {len(documentation.instrument_tags)}")
    print(f"I/O Points: {len(documentation.io_points)}")
    print(f"Control Loops: {len(documentation.control_loops)}")
    print(f"Cables: {len(documentation.cable_schedule)}")
    print(f"Total Power: {documentation.power_requirements.get('total_power_consumption', 0)} W")
    
    # Export to Excel
    formatter = DocumentFormatter()
    output_path = "/tmp/control_system_documentation.xlsx"
    formatter.export_to_excel(documentation, output_path)
    print(f"Documentation exported to: {output_path}")
    
    return documentation

if __name__ == "__main__":
    asyncio.run(test_documentation_generation())